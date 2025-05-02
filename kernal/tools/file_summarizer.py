# ./tools/file_summarizer.py
# python -m tools.file_summarizer
# Summary: A tool that takes a list of file paths and summarizes their content. For DOCX, PDF, and XLSX
# files it extracts text using methods from the drive_file_scanner. For all other files it reads the file normally.
# It then uses the ChatGPT API to obtain a brief summary and writes the results to an output file.

import os
import concurrent.futures
from docx import Document
import PyPDF2
from openpyxl import load_workbook
import time

from tools.chat_gpt_api import ChatGptApi
from globals.KernalBasics import *

logger = get_logger(__name__)


class FileSummarizer:
    def __init__(self, file_paths, output_file="file_summary.txt", model="gpt-4o-mini"):
        self.file_paths = file_paths
        self.output_file = output_file
        self.model = model
        self.gpt = ChatGptApi()

    def estimate_tokens(self, text):
        """Estimate token count based on character count (approx 4 chars per token)"""
        if not text:
            return 0
        return len(text) // 4

    def read_docx(self, file_path):
        try:
            doc = Document(file_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            line_count = len(text.split('\n'))
            return text, line_count
        except Exception as e:
            logger.error(f"Error reading DOCX {file_path}: {e}")
            return "", 0

    def read_pdf(self, file_path):
        content = []
        try:
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    content.append(page.extract_text())
            text = "\n".join(content)
            line_count = len(text.split('\n'))
            return text, line_count
        except Exception as e:
            logger.error(f"Error reading PDF {file_path}: {e}")
            return "", 0

    def read_xlsx(self, file_path):
        content = []
        try:
            wb = load_workbook(filename=file_path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    content.append(
                        " | ".join(
                            [str(cell) if cell is not None else "" for cell in row]
                        )
                    )
            text = "\n".join(content)
            line_count = len(content)
            return text, line_count
        except Exception as e:
            logger.error(f"Error reading XLSX {file_path}: {e}")
            return "", 0

    def extract_content(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        # If the file is one of the supported types, use DriveFileScanner reader methods.
        if ext == ".docx":
            try:
                # Create a temporary instance using the file's directory (db not used here)
                return self.read_docx(file_path)
            except Exception as e:
                logger.error(f"Error extracting DOCX file {file_path}: {e}")
                return "", 0
        elif ext == ".pdf":
            try:
                return self.read_pdf(file_path)
            except Exception as e:
                logger.error(f"Error extracting PDF file {file_path}: {e}")
                return "", 0
        elif ext == ".xlsx":
            try:
                return self.read_xlsx(file_path)
            except Exception as e:
                logger.error(f"Error extracting XLSX file {file_path}: {e}")
                return "", 0
        else:
            # For all other files, open as text.
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()
                    line_count = len(text.splitlines())
                    return text, line_count
            except Exception as e:
                logger.error(f"Error reading file {file_path}: {e}")
                return "", 0

    def build_prompt(self, file_path, content):
        prompt = (
            "Compress the following file content into a brief summary of less than 200 words in german.\n"
            "Please provide a clear and concise summary that covers:\n"
            "- The key topics and main ideas.\n"
            "- Important details or events.\n"
            "- For code files: a brief explanation of the code's functionality, "
            "purpose, and any significant logic or structure.\n"
            "- For meeting protocols or logs: the primary actions, dates, and any outcomes or follow-up items.\n"
            "- For other documents: a high-level overview that captures the essential points.\n"
            "Ensure that your summary is objective, well-structured (using bullet points or short paragraphs).\n"
            "Also include what kind of document it could be out of the following:\n"
            "- Meeting protocol\n"
            "- Code file\n"
            "- Documentation\n"
            "- Other\n\n"
            f"File: {file_path}\n\n{content}\n\n"
            "You response as json-format with key-value pairs like this:\n\n"
            "Summary: your german sumamry here\n"
            "Document type: your document type here\n"
            "Meta: any additional notes here for example the date\n"
            ""
        )
        return prompt

    def summarize_file(self, file_path):
        logger.info(f"Processing file: {file_path}")
        content, line_count = self.extract_content(file_path)
        token_count = self.estimate_tokens(content)
        
        try:
            entry_creation_date = time.ctime(os.path.getmtime(file_path))
        except Exception as e:
            logger.error(f"Error getting creation date for {file_path}: {e}")
            entry_creation_date = "N/A"
            
        if not content:
            return f"Modified at {entry_creation_date} | Lines: 0 | Tokens: 0:\n" + f"Could not extract content from {file_path}"
            
        prompt = self.build_prompt(file_path, content)
        try:
            summary = self.gpt.prompt(self.model, prompt, 2000)
        except Exception as e:
            summary = f"Error summarizing file {file_path}: {e}"
            
        return f"Modified at {entry_creation_date} | Lines: {line_count} | Tokens: {token_count}:\n" + summary

    def process(self):
        results = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=100) as executor:
            future_to_path = {
                executor.submit(self.summarize_file, path): path
                for path in self.file_paths
            }
            for future in concurrent.futures.as_completed(future_to_path):
                path = future_to_path[future]
                try:
                    summary = future.result()
                except Exception as exc:
                    summary = f"Error processing file {path}: {exc}"
                results[path] = summary
                logger.info(f"Summarized {path}")
        return results


if __name__ == "__main__":
    file_paths = [
        "/Users/thurefoken/Desktop/Repos/digitalermitarbeiter/agent_kernal.py",
        "/Users/thurefoken/Desktop/Repos/digitalermitarbeiter/base_prompt.txt",
    ]
    summarizer = FileSummarizer(file_paths)
    results = summarizer.process()
    logger.info(results)
